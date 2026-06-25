import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AZUL = "1A56DB"
AZUL_CLARO = "E8F0FE"
CINZA = "F3F4F6"


def gerar_excel(titulo: str, headers: list, rows: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    # Título
    ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=14, color=AZUL)
    ws.cell(row=2, column=1, value=f"Gerado em {date.today().strftime('%d/%m/%Y')}").font = Font(
        size=9, color="6B7280"
    )
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 14

    # Cabeçalho (linha 4)
    header_fill = PatternFill(fill_type="solid", fgColor=AZUL)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_border = Border(
        bottom=Side(style="thin", color="FFFFFF")
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = header_border
    ws.row_dimensions[4].height = 18

    # Dados (a partir da linha 5)
    for row_idx, row in enumerate(rows, 5):
        fill_color = CINZA if (row_idx - 5) % 2 == 0 else "FFFFFF"
        fill = PatternFill(fill_type="solid", fgColor=fill_color)
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_idx].height = 15

    # Largura automática das colunas
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row in rows:
            val = str(row[col_idx - 1]) if col_idx - 1 < len(row) else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
